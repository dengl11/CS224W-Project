import openpyxl
from collections import defaultdict
import pickle as pkl

# def loadData(filename):
# 	wb = openpyxl.load_workbook(filename)
# 	active = wb.get_sheet_by_name("globalterrorismdb_0617dist")
# 	return active

# def iter_rows(active):
# 	for row in active.iter_rows():
# 		yield [cell.value for cell in row]

# def printDict(active):
# 	dataList = []
# 	res = iter_rows(active)
# 	keys = next(res)
# 	for new in res:
# 		dataList.append(dict(zip(keys,new)))
# 	return dataList

# def processType(data):
# 	newdict = {}
# 	for row in data:
# 		attack_id = row['eventid']
# 		newrow = {}
# 		for key in row:
# 			key = key.encode('ascii', 'ignore')
# 			if key == 'latitude' or key == 'longitude':
# 				if row[key]:
# 					newrow[key] = float(str(row[key]))
# 					continue
# 			if row[key]:
# 				try:
# 					newint = int(float(row[key]))
# 					newrow[key] = newint
# 				except (AttributeError, TypeError, ValueError):
# 					try:
# 						newrow[key] = row[key].encode('ascii', 'ignore')
# 					except (AttributeError, TypeError, ValueError):
# 						newrow[key] = row[key]
# 			elif row[key] is not None:
# 				if str(row[key]) == '0L' or str(row[key]) == '0':
# 					newrow[key] = 0
# 			else:
# 				newrow[key] = row[key]
# 		newdict[attack_id] = newrow
# 	return newdict


# def main():
# 	filename = "../../data/excel/globalterrorismdb_0617dist.xlsx"
# 	active = loadData(filename)
# 	iter_rows(active)
# 	data = printDict(active)
# 	data = processType(data)
# 	pkl.dump(data, open("../../data/pkl/GTD_dict.p", "wb"))
def main():
	with open("../../data/pkl/GTD_raw.p", "rb") as f:
		terror_attack_data_raw = pkl.load(f)

	terror_dict = {}
	for attack in terror_attack_data_raw:
		key = attack['eventid']
		terror_dict[key] = attack

	pkl.dump(terror_dict, open("../../data/pkl/GTD_dict.p", "wb"))

main()