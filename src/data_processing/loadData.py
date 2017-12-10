import openpyxl
from collections import defaultdict
import pickle as pkl

def loadData(filename):
	wb = openpyxl.load_workbook(filename + "xlsm")
	active = wb.get_sheet_by_name(filename)
	return active

def iter_rows(active):
	for row in active.iter_rows():
		yield [cell.value for cell in row]

def printDict(active):
	dataList = []
	res = iter_rows(active)
	keys = next(res)
	for new in res:
		dataList.append(dict(zip(keys,new)))
	return dataList

def processType(data):
	newList = []
	for row in data:
		newrow = {}
		for key in row:
			key = str(key)
			if key == "group" or key == "cowmastercountry":
				newrow[key] = str(row[key])
			else:
				newrow[key] = int(row[key])
		newList.append(newrow)
	return newList

def main():
	filename = "../../data/excel/lethality_paper__Dataset_Jan15."
	active = loadData(filename)
	iter_rows(active)
	data = printDict(active)
	data = processType(data)
	pkl.dump(data, open("../../data/pkl/baad.p", "wb"))

main()